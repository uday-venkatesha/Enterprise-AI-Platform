import io

from openpyxl import load_workbook


def extract_text_from_xlsx(raw_bytes: bytes) -> str:
    # data_only=True gives us computed VALUES of formulas, not the formula text
    # ("=SUM(A1:A5)" -> the actual number). read_only=True is a memory-efficient
    # mode for just reading, which matters for big spreadsheets.
    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)

    lines: list[str] = []
    # A workbook has multiple SHEETS (tabs).
    for sheet in workbook.worksheets:
        # Label each sheet so the AI later knows which tab data came from.
        lines.append(f"# Sheet: {sheet.title}")

        # iter_rows(values_only=True) yields each row as a tuple of cell values.
        for row in sheet.iter_rows(values_only=True):
            # Turn the row's cells into strings, skipping empty (None) cells.
            cells = [str(value) for value in row if value is not None]
            # Only keep rows that actually had content.
            if cells:
                # Join a row's cells with " | " so the AI can see column structure
                # as readable text. This is the KEY design choice — see below.
                lines.append(" | ".join(cells))

    return "\n".join(lines).strip()